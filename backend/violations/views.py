import csv
import io
import mimetypes
from calendar import monthrange
from datetime import datetime, timedelta
from pathlib import Path

from django.db.models import Count, Q
from django.http import FileResponse, HttpResponse
from django.utils import timezone
from rest_framework import filters, serializers, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend
import django_filters
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import (
    Image as RLImage, SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    HRFlowable, KeepTogether,
)
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.graphics import renderPDF
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.units import pixels_to_EMU

from cameras.models import SystemSettings
from .models import Violation
from .serializers import (
    ViolationSerializer,
    ViolationSummarySerializer,
    ViolationWeeklyChartSerializer,
)
from users.models import AdminNotification, UserNotification
from users.permissions import (
    CanAccessViolationRecords,
    CanViewViolationAnalytics,
    CanViewViolations,
    IsAdmin,
    IsYoloService,
    has_user_permission,
    is_admin_user,
)


def _parse_query_date(value, fmt):
    if not value:
        return None

    try:
        return datetime.strptime(value, fmt).date()
    except (TypeError, ValueError):
        return None


def _apply_date_filters(queryset, params):
    today = timezone.localdate()
    date_filter = params.get('date')
    specific_date = _parse_query_date(params.get('specific_date'), '%Y-%m-%d')
    specific_month = params.get('specific_month')
    year = params.get('year')
    date_from = _parse_query_date(params.get('date_from'), '%Y-%m-%d')
    date_to = _parse_query_date(params.get('date_to'), '%Y-%m-%d')

    if year:
        try:
            queryset = queryset.filter(detected_at__year=int(year))
        except (TypeError, ValueError):
            pass

    if date_filter == 'today':
        queryset = queryset.filter(detected_at__date=today)
    elif date_filter == 'week':
        queryset = queryset.filter(detected_at__date__gte=today - timedelta(days=6))
    elif date_filter == 'month':
        queryset = queryset.filter(detected_at__date__gte=today.replace(day=1))
    elif date_filter:
        parsed_date = _parse_query_date(date_filter, '%Y-%m-%d')
        if parsed_date:
            queryset = queryset.filter(detected_at__date=parsed_date)

    if specific_date:
        queryset = queryset.filter(detected_at__date=specific_date)

    if specific_month:
        parsed_month = _parse_query_date(f'{specific_month}-01', '%Y-%m-%d')
        if parsed_month:
            queryset = queryset.filter(
                detected_at__year=parsed_month.year,
                detected_at__month=parsed_month.month,
            )

    if date_from:
        queryset = queryset.filter(detected_at__date__gte=date_from)
    if date_to:
        queryset = queryset.filter(detected_at__date__lte=date_to)

    return queryset


def _apply_violation_filters(queryset, params):
    queryset = _apply_date_filters(queryset, params)

    location = params.get('location') or params.get('camera')
    if location:
        if str(location).isdigit():
            queryset = queryset.filter(camera_id=location)
        else:
            queryset = queryset.filter(
                Q(camera__name__icontains=location) |
                Q(camera__location__icontains=location)
            )

    status = params.get('detection_status') or params.get('status')
    if status:
        queryset = queryset.filter(detection_status=status)

    review_status = params.get('review_status')
    if review_status:
        queryset = queryset.filter(review_status=review_status)

    classification = params.get('classification')
    if classification:
        queryset = queryset.filter(classification=classification)

    return queryset


def _resolve_chart_date_range(params):
    today = timezone.localdate()
    date_filter = params.get('date')
    specific_date = _parse_query_date(params.get('specific_date'), '%Y-%m-%d')
    specific_month = params.get('specific_month')
    date_from = _parse_query_date(params.get('date_from'), '%Y-%m-%d')
    date_to = _parse_query_date(params.get('date_to'), '%Y-%m-%d')

    if date_from or date_to:
        start_date = date_from or date_to
        end_date = date_to or date_from
        if start_date and end_date and start_date > end_date:
            start_date, end_date = end_date, start_date
        return start_date, end_date

    if specific_date:
        return specific_date, specific_date

    if specific_month:
        parsed_month = _parse_query_date(f'{specific_month}-01', '%Y-%m-%d')
        if parsed_month:
            last_day = monthrange(parsed_month.year, parsed_month.month)[1]
            return parsed_month, parsed_month.replace(day=last_day)

    if date_filter == 'today':
        return today, today
    if date_filter == 'week':
        return today - timedelta(days=6), today
    if date_filter == 'month':
        return today.replace(day=1), today

    parsed_date = _parse_query_date(date_filter, '%Y-%m-%d')
    if parsed_date:
        return parsed_date, parsed_date

    return today - timedelta(days=6), today


class ViolationFilter(django_filters.FilterSet):
    detected_at__gte = django_filters.IsoDateTimeFilter(field_name='detected_at', lookup_expr='gte')
    detected_at__lte = django_filters.IsoDateTimeFilter(field_name='detected_at', lookup_expr='lte')
    
    year = django_filters.NumberFilter(field_name='detected_at', lookup_expr='year')

    class Meta:
        model = Violation
        fields = ['detection_status', 'classification', 'camera', 'review_status', 'year']


class ViolationViewSet(viewsets.ModelViewSet):
    queryset = Violation.objects.select_related('reviewed_by__profile', 'plate_corrected_by').all().order_by('-detected_at')
    serializer_class = ViolationSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = ViolationFilter
    ordering_fields = ['detected_at', 'confidence_score']
    ordering = ['-detected_at']

    def get_queryset(self):
        qs = super().get_queryset()
        search = (self.request.query_params.get('search') or '').strip()
        qs = _apply_violation_filters(qs, self.request.query_params)

        if search:
            if not is_admin_user(self.request.user):
                if len(search) != 3:
                    return qs.none()
            qs = qs.filter(plate_number__icontains=search)
            if not getattr(self.request, "_plate_search_activity_logged", False):
                AdminNotification.create_for_plate_search(
                    actor=self.request.user,
                    search_term=search,
                )
                self.request._plate_search_activity_logged = True

        return qs

    def get_permissions(self):
        if self.action == 'create':
            return [IsYoloService()]
        if self.action == 'list':
            return [CanAccessViolationRecords()]
        if self.action in ['retrieve', 'evidence', 'update', 'partial_update', 'correct_plate']:
            return [CanViewViolations()]
        return [IsAdmin()]

    def perform_create(self, serializer):
        violation = serializer.save()
        settings_obj = SystemSettings.get_settings()
        if settings_obj.notify_on_new_detection:
            AdminNotification.create_for_new_detection(violation=violation)
            UserNotification.create_for_new_detection(violation=violation)

    def perform_update(self, serializer):
        if not has_user_permission(self.request.user, "can_update_violation_status"):
            raise PermissionDenied("You do not have permission to update violation status.")

        instance = self.get_object()
        previous_status = instance.review_status
        new_status = serializer.validated_data.get('review_status')

        if new_status and new_status != instance.review_status and new_status in ('reviewed', 'resolved'):
            updated_violation = serializer.save(reviewed_by=self.request.user, reviewed_at=timezone.now())
        elif new_status == 'pending':
            updated_violation = serializer.save(reviewed_by=None, reviewed_at=None)
        else:
            updated_violation = serializer.save()

        if new_status and new_status != previous_status:
            AdminNotification.create_for_violation_action(
                actor=self.request.user,
                violation=updated_violation,
                previous_status=instance.get_review_status_display(),
                new_status=updated_violation.get_review_status_display(),
            )

    @action(detail=True, methods=['patch'], url_path='correct-plate')
    def correct_plate(self, request, pk=None):
        if not (request.user.is_staff or getattr(getattr(request.user, 'profile', None), 'role', None) == 'admin' or has_user_permission(request.user, "can_correct_plate_number")):
            raise PermissionDenied("You do not have permission to correct plate numbers.")

        plate_number_corrected = str(request.data.get('plate_number_corrected', '')).strip()
        if not plate_number_corrected:
            raise serializers.ValidationError({
                'plate_number_corrected': 'Plate number corrected must not be empty.',
            })
        if len(plate_number_corrected) > 20:
            raise serializers.ValidationError({
                'plate_number_corrected': 'Ensure this field has no more than 20 characters.',
            })

        violation = self.get_object()
        original_display_plate = violation.plate_number_corrected or violation.plate_number or 'N/A'

        violation.plate_number_corrected = plate_number_corrected
        violation.plate_corrected_by = request.user
        violation.plate_corrected_at = timezone.now()
        violation.save(update_fields=['plate_number_corrected', 'plate_corrected_by', 'plate_corrected_at'])

        # Notify Admin if operator corrected plate
        AdminNotification.create_for_plate_correction(
            actor=request.user,
            violation=violation,
            original_plate=original_display_plate,
            corrected_plate=plate_number_corrected,
        )

        # Notify Operators when plate is corrected
        UserNotification.create_for_plate_correction(
            actor=request.user,
            violation=violation,
            original_plate=original_display_plate,
            corrected_plate=plate_number_corrected,
        )

        return Response(self.get_serializer(violation).data)

    @action(detail=True, methods=['get'], url_path='evidence')
    def evidence(self, request, pk=None):
        violation = self.get_object()
        variant = str(request.query_params.get("variant", "")).lower()
        image_field = violation.plate_crop_image if variant == 'plate' else violation.evidence_image

        if not image_field:
            return Response({"error": "Evidence image not found."}, status=404)

        download_requested = str(request.query_params.get("download", "")).lower() in {
            "1", "true", "yes",
        }

        AdminNotification.create_for_evidence_view(
            actor=request.user,
            violation=violation,
        )

        evidence_name = Path(image_field.name).name
        content_type, _ = mimetypes.guess_type(evidence_name)
        response = FileResponse(
            image_field.open('rb'),
            as_attachment=download_requested,
            filename=evidence_name,
            content_type=content_type or 'application/octet-stream',
        )
        response["Cache-Control"] = "private, no-store"
        response["X-Content-Type-Options"] = "nosniff"
        return response


class ViolationSummaryView(APIView):
    permission_classes = [CanViewViolationAnalytics]

    def get(self, request):
        today = timezone.localdate()
        week_start = today - timedelta(days=6)
        queryset = _apply_violation_filters(
            Violation.objects.select_related("camera"),
            request.query_params,
        )

        class_counts = {
            row["classification"]: row["count"]
            for row in queryset.values("classification").annotate(count=Count("id"))
        }
        camera_counts = queryset.values("camera__name").annotate(count=Count("id")).order_by("-count", "camera__name")

        summary = {
            "total_violations": queryset.count(),
            "pending_violations": queryset.filter(review_status="pending").count(),
            "reviewed_violations": queryset.filter(review_status="reviewed").count(),
            "resolved_violations": queryset.filter(review_status="resolved").count(),
            "today_violations": queryset.filter(detected_at__date=today).count(),
            "this_week_violations": queryset.filter(detected_at__date__range=(week_start, today)).count(),
            "by_class": [
                {
                    "classification": code,
                    "label": label,
                    "count": class_counts.get(code, 0),
                }
                for code, label in Violation.CLASSIFICATION_CHOICES
            ],
            "by_camera": [
                {
                    "camera_name": row["camera__name"] or "Unknown",
                    "count": row["count"],
                }
                for row in camera_counts
            ],
        }

        serializer = ViolationSummarySerializer(summary)
        return Response(serializer.data)


class ViolationWeeklyChartView(APIView):
    permission_classes = [CanViewViolationAnalytics]

    def get(self, request):
        start_date, end_date = _resolve_chart_date_range(request.query_params)
        queryset = _apply_violation_filters(Violation.objects.all(), request.query_params).filter(
            detected_at__date__range=(start_date, end_date)
        )

        counts_by_day = {
            row["detected_at__date"]: row["count"]
            for row in (
                queryset
                .values("detected_at__date")
                .annotate(count=Count("id"))
            )
        }

        payload = [
            {
                "date": start_date + timedelta(days=offset),
                "count": counts_by_day.get(start_date + timedelta(days=offset), 0),
            }
            for offset in range((end_date - start_date).days + 1)
        ]

        serializer = ViolationWeeklyChartSerializer(payload, many=True)
        return Response(serializer.data)


class ViolationExportView(APIView):
    permission_classes = [IsAuthenticated]

    @staticmethod
    def _get_tmc_logo_path():
        logo_path = Path(__file__).resolve().parents[2] / 'frontend' / 'public' / 'tmc.jpg'
        return logo_path if logo_path.exists() else None

    @staticmethod
    def _get_display_plate_number(violation):
        return violation.plate_number_corrected or violation.plate_number or 'N/A'

    @staticmethod
    def _is_plate_corrected(violation):
        return bool((violation.plate_number_corrected or '').strip())

    def _get_qs(self, request):
        qs = _apply_violation_filters(
            Violation.objects.all().order_by('-detected_at'),
            request.query_params,
        )
        search = (request.query_params.get('search') or '').strip()

        if search:
            if not is_admin_user(request.user) and len(search) != 3:
                return qs.none()
            qs = qs.filter(plate_number__icontains=search)
        return qs

    def get(self, request):
        if not has_user_permission(request.user, "can_view_reports"):
            raise PermissionDenied("You do not have permission to view reports.")

        if not has_user_permission(request.user, "can_export_reports"):
            raise PermissionDenied("You do not have permission to export reports.")

        fmt = request.query_params.get('export_format', 'xlsx').lower()
        qs = self._get_qs(request)
        AdminNotification.create_for_report_export(
            actor=request.user,
            export_format=fmt,
            record_count=qs.count(),
        )

        if fmt == 'pdf':
            return self._pdf(qs)
        if fmt == 'xlsx':
            return self._xlsx(qs)
        return self._csv(qs)

    def _csv(self, qs):
        now = datetime.now()
        logo_path = self._get_tmc_logo_path()
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = (
            f'attachment; filename="saferide_violations_'
            f'{now.strftime("%Y%m%d_%H%M%S")}.csv"'
        )

        # BOM for Excel UTF-8 recognition
        response.write('\ufeff')
        w = csv.writer(response)
        w.writerow(['Traffic Management Center (TMC)'])
        if logo_path:
            w.writerow([f'TMC Logo File: {logo_path.name}'])

        # ── report header rows ───────────────────────────────────
        w.writerow(['SafeRide — Violation Report'])
        w.writerow([f'Generated: {now.strftime("%B %d, %Y %I:%M %p")}'])
        w.writerow([f'Total Records: {qs.count()}'])
        w.writerow([])  # blank separator

        # ── summary stats ────────────────────────────────────────
        violations_count = qs.filter(detection_status='violation').count()
        reviewed_count = qs.filter(review_status='reviewed').count()
        resolved_count = qs.filter(review_status='resolved').count()
        pending_count = (
            qs.filter(review_status='pending').count()
            + qs.filter(review_status__isnull=True).count()
        )

        w.writerow(['Summary'])
        w.writerow(['Violations', 'Reviewed', 'Resolved', 'Pending'])
        w.writerow([violations_count, reviewed_count, resolved_count, pending_count])
        w.writerow([])  # blank separator

        # ── column headers ───────────────────────────────────────
        w.writerow([
            '#', 'ID', 'Date', 'Time', 'Camera', 'Classification',
            'Plate Number', 'Plate Corrected', 'Detection Status', 'Confidence (%)',
            'Review Status',
        ])

        # ── data rows ───────────────────────────────────────────
        for idx, v in enumerate(qs, 1):
            w.writerow([
                idx,
                v.id,
                v.detected_at.strftime('%Y-%m-%d'),
                v.detected_at.strftime('%H:%M:%S'),
                v.camera.name if v.camera else 'Unknown',
                v.get_classification_display(),
                self._get_display_plate_number(v),
                'Yes' if self._is_plate_corrected(v) else 'No',
                v.get_detection_status_display(),
                f'{v.confidence_score * 100:.1f}%',
                v.get_review_status_display(),
            ])

        # ── footer ───────────────────────────────────────────────
        w.writerow([])
        w.writerow([f'End of Report — {qs.count()} records exported'])

        return response

    def _xlsx(self, qs):
        now = datetime.now()
        logo_path = self._get_tmc_logo_path()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Violation Report'
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = 'A13'

        border_color = 'E2E8F0'
        accent_color = '3B82F6'
        brand_color = '1E293B'
        muted_color = '64748B'
        light_fill = 'F8FAFC'
        header_fill = 'EEF4FF'
        thin_side = Side(style='thin', color=border_color)
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        column_widths = {
            'A': 8,
            'B': 12,
            'C': 14,
            'D': 12,
            'E': 18,
            'F': 18,
            'G': 16,
            'H': 16,
            'I': 15,
            'J': 16,
        }
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        def _column_width_to_pixels(width):
            # Approximate Excel column width conversion for image placement.
            return int(width * 7 + 5) if width >= 1 else int(width * 12)

        if logo_path:
            logo = XLImage(str(logo_path))
            max_width = 120
            scale = max_width / logo.width
            logo.width = max_width
            logo.height = int(logo.height * scale)
            header_columns = list(column_widths.keys())
            header_width_pixels = [_column_width_to_pixels(column_widths[column]) for column in header_columns]
            start_x_pixels = max(0, (sum(header_width_pixels) - logo.width) // 2)
            anchor_col = 0
            remaining_offset = start_x_pixels
            while anchor_col < len(header_width_pixels) - 1 and remaining_offset >= header_width_pixels[anchor_col]:
                remaining_offset -= header_width_pixels[anchor_col]
                anchor_col += 1

            logo.anchor = OneCellAnchor(
                _from=AnchorMarker(
                    col=anchor_col,
                    row=0,
                    colOff=pixels_to_EMU(remaining_offset),
                    rowOff=pixels_to_EMU(4),
                ),
                ext=XDRPositiveSize2D(
                    cx=pixels_to_EMU(logo.width),
                    cy=pixels_to_EMU(logo.height),
                ),
            )
            worksheet.add_image(logo)

        worksheet.row_dimensions[1].height = 70
        worksheet.row_dimensions[5].height = 22
        worksheet.row_dimensions[6].height = 26
        worksheet.row_dimensions[7].height = 20
        worksheet.row_dimensions[8].height = 20

        worksheet.merge_cells('A5:J5')
        worksheet['A5'] = 'Traffic Management Center (TMC)'
        worksheet['A5'].font = Font(name='Calibri', size=12, bold=True, color=accent_color)
        worksheet['A5'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.merge_cells('A6:J6')
        worksheet['A6'] = 'SafeRide Violation Report'
        worksheet['A6'].font = Font(name='Calibri', size=18, bold=True, color=brand_color)
        worksheet['A6'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.merge_cells('A7:J7')
        worksheet['A7'] = f'Generated: {now.strftime("%B %d, %Y %I:%M %p")}'
        worksheet['A7'].font = Font(name='Calibri', size=10, color=muted_color)
        worksheet['A7'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.merge_cells('A8:J8')
        worksheet['A8'] = f'Total Records: {qs.count()}'
        worksheet['A8'].font = Font(name='Calibri', size=10, color=muted_color)
        worksheet['A8'].alignment = Alignment(horizontal='center', vertical='center')

        violations_count = qs.filter(detection_status='violation').count()
        reviewed_count = qs.filter(review_status='reviewed').count()
        resolved_count = qs.filter(review_status='resolved').count()
        pending_count = (
            qs.filter(review_status='pending').count()
            + qs.filter(review_status__isnull=True).count()
        )

        worksheet['A10'] = 'Summary'
        worksheet['A10'].font = Font(name='Calibri', size=11, bold=True, color=brand_color)

        summary_headers = ['Violations', 'Reviewed', 'Resolved', 'Pending']
        summary_values = [violations_count, reviewed_count, resolved_count, pending_count]
        for index, (label, value) in enumerate(zip(summary_headers, summary_values), start=1):
            cell = worksheet.cell(row=11, column=index, value=label)
            cell.font = Font(name='Calibri', size=10, bold=True, color=brand_color)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(fill_type='solid', fgColor=header_fill)
            cell.border = thin_border

            value_cell = worksheet.cell(row=12, column=index, value=value)
            value_cell.font = Font(name='Calibri', size=11, bold=True, color=brand_color)
            value_cell.alignment = Alignment(horizontal='center')
            value_cell.fill = PatternFill(fill_type='solid', fgColor=light_fill)
            value_cell.border = thin_border

        headers = [
            '#', 'ID', 'Date', 'Time', 'Camera', 'Classification',
            'Plate Number', 'Detection Status', 'Confidence (%)', 'Review Status',
        ]

        header_row = 14
        for column_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=header_row, column=column_index, value=header)
            cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(fill_type='solid', fgColor=brand_color)
            cell.border = thin_border

        for row_index, violation in enumerate(qs, start=15):
            values = [
                row_index - 14,
                violation.id,
                violation.detected_at.strftime('%Y-%m-%d'),
                violation.detected_at.strftime('%H:%M:%S'),
                violation.camera.name if violation.camera else 'Unknown',
                violation.get_classification_display(),
                self._get_display_plate_number(violation),
                violation.get_detection_status_display(),
                float(f'{violation.confidence_score * 100:.1f}'),
                violation.get_review_status_display(),
            ]

            for column_index, value in enumerate(values, start=1):
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                if row_index % 2 == 0:
                    cell.fill = PatternFill(fill_type='solid', fgColor=light_fill)

            worksheet.cell(row=row_index, column=9).number_format = '0.0"%"'

        worksheet.auto_filter.ref = f'A{header_row}:J{max(header_row, qs.count() + 14)}'

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="saferide_violations_'
            f'{now.strftime("%Y%m%d_%H%M%S")}.xlsx"'
        )
        return response

    def _pdf(self, qs):
        buffer = io.BytesIO()
        page_w, page_h = landscape(A4)
        doc = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            rightMargin=0.5 * inch, leftMargin=0.5 * inch,
            topMargin=0.5 * inch, bottomMargin=0.7 * inch,
        )

        # ── colours ──────────────────────────────────────────────
        BRAND      = colors.HexColor('#1e293b')   # slate-800
        ACCENT     = colors.HexColor('#3b82f6')   # blue-500
        LIGHT_BG   = colors.HexColor('#f8fafc')   # slate-50
        BORDER     = colors.HexColor('#e2e8f0')   # slate-200
        TEXT_DARK  = colors.HexColor('#1e293b')
        TEXT_MUTED = colors.HexColor('#64748b')   # slate-500
        SUCCESS    = colors.HexColor('#16a34a')
        WARNING    = colors.HexColor('#f59e0b')
        DANGER     = colors.HexColor('#ef4444')

        # ── styles ───────────────────────────────────────────────
        styles = getSampleStyleSheet()
        s_title = ParagraphStyle(
            'RTitle', parent=styles['Title'],
            fontName='Helvetica-Bold', fontSize=20,
            textColor=BRAND, spaceAfter=2,
        )
        s_subtitle = ParagraphStyle(
            'RSub', parent=styles['Normal'],
            fontName='Helvetica', fontSize=9,
            textColor=TEXT_MUTED, spaceAfter=6,
        )
        s_section = ParagraphStyle(
            'RSection', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=11,
            textColor=BRAND, spaceBefore=14, spaceAfter=6,
        )
        s_cell = ParagraphStyle(
            'RCell', parent=styles['Normal'],
            fontName='Helvetica', fontSize=7.5,
            textColor=TEXT_DARK, leading=10,
        )
        s_cell_center = ParagraphStyle(
            'RCellC', parent=s_cell, alignment=TA_CENTER,
        )
        s_header_cell = ParagraphStyle(
            'RHCell', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=7.5,
            textColor=colors.white, alignment=TA_CENTER, leading=10,
        )

        elements = []
        logo_path = self._get_tmc_logo_path()

        # ── page footer (page number) ───────────────────────────
        def _footer(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 7)
            canvas.setFillColor(TEXT_MUTED)
            canvas.drawRightString(
                page_w - 0.5 * inch, 0.35 * inch,
                f'Page {doc.page}',
            )
            canvas.drawString(
                0.5 * inch, 0.35 * inch,
                'SafeRide Violation Management System — Confidential',
            )
            # thin line above footer
            canvas.setStrokeColor(BORDER)
            canvas.setLineWidth(0.5)
            canvas.line(0.5 * inch, 0.5 * inch, page_w - 0.5 * inch, 0.5 * inch)
            canvas.restoreState()

        # ── header section ───────────────────────────────────────
        now = datetime.now()
        total = qs.count()

        if logo_path:
            logo_reader = ImageReader(str(logo_path))
            logo_w, logo_h = logo_reader.getSize()
            display_width = 1.35 * inch
            display_height = display_width * (logo_h / logo_w)
            logo = RLImage(str(logo_path), width=display_width, height=display_height)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 6))

        elements.append(Paragraph(
            'Traffic Management Center (TMC)',
            ParagraphStyle(
                'RTmc',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=11,
                textColor=ACCENT,
                alignment=TA_CENTER,
                spaceAfter=8,
            )
        ))
        elements.append(Paragraph('SafeRide', s_title))
        elements.append(Paragraph('Violation Report', ParagraphStyle(
            'RSubtitle2', parent=styles['Normal'],
            fontName='Helvetica', fontSize=13,
            textColor=ACCENT, spaceAfter=4,
        )))
        elements.append(HRFlowable(
            width='100%', thickness=1.5, color=ACCENT,
            spaceAfter=10, spaceBefore=2,
        ))

        # meta row
        meta_text = (
            f'<font color="#64748b">'
            f'<b>Generated:</b> {now.strftime("%B %d, %Y  %I:%M %p")} &nbsp;&nbsp;|&nbsp;&nbsp; '
            f'<b>Total Records:</b> {total}'
            f'</font>'
        )
        elements.append(Paragraph(meta_text, s_subtitle))
        elements.append(Spacer(1, 6))

        # ── summary cards ────────────────────────────────────────
        violations_count = qs.filter(detection_status='violation').count()
        reviewed_count = qs.filter(review_status='reviewed').count()
        resolved_count = qs.filter(review_status='resolved').count()
        pending_count = qs.filter(review_status='pending').count() + qs.filter(review_status__isnull=True).count()

        card_data = [
            [
                Paragraph(f'<font size="14"><b>{violations_count}</b></font><br/>'
                          f'<font size="7" color="#64748b">Violations</font>', s_cell_center),
                Paragraph(f'<font size="14"><b>{reviewed_count}</b></font><br/>'
                          f'<font size="7" color="#64748b">Reviewed</font>', s_cell_center),
                Paragraph(f'<font size="14"><b>{resolved_count}</b></font><br/>'
                          f'<font size="7" color="#64748b">Resolved</font>', s_cell_center),
                Paragraph(f'<font size="14"><b>{pending_count}</b></font><br/>'
                          f'<font size="7" color="#64748b">Pending</font>', s_cell_center),
            ]
        ]
        card_w = (page_w - 1.0 * inch) / 4
        card_table = Table(card_data, colWidths=[card_w] * 4, rowHeights=[48])
        card_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#fef2f2')),   # red-50
            ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#eff6ff')),   # blue-50
            ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#f0fdf4')),   # green-50
            ('BACKGROUND', (3, 0), (3, 0), colors.HexColor('#fffbeb')),   # amber-50
            ('BOX', (0, 0), (0, 0), 0.75, colors.HexColor('#fecaca')),
            ('BOX', (1, 0), (1, 0), 0.75, colors.HexColor('#bfdbfe')),
            ('BOX', (2, 0), (2, 0), 0.75, colors.HexColor('#bbf7d0')),
            ('BOX', (3, 0), (3, 0), 0.75, colors.HexColor('#fde68a')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('ROUNDEDCORNERS', [6, 6, 6, 6]),
        ]))
        elements.append(card_table)
        elements.append(Spacer(1, 14))

        # ── section label ────────────────────────────────────────
        elements.append(Paragraph('Violation Details', s_section))

        # ── data table ───────────────────────────────────────────
        headers = ['#', 'ID', 'Date', 'Time', 'Camera', 'Classification',
                   'Plate', 'Status', 'Confidence', 'Review']
        header_row = [Paragraph(h, s_header_cell) for h in headers]
        data = [header_row]

        def _status_color(status):
            s = (status or '').lower()
            if s == 'resolved':
                return f'<font color="#16a34a"><b>{status}</b></font>'
            elif s == 'reviewed':
                return f'<font color="#3b82f6"><b>{status}</b></font>'
            return f'<font color="#f59e0b"><b>{status}</b></font>'

        def _conf_color(score):
            pct = score * 100
            label = f'{pct:.1f}%'
            if pct >= 75:
                return f'<font color="#16a34a"><b>{label}</b></font>'
            elif pct >= 50:
                return f'<font color="#f59e0b"><b>{label}</b></font>'
            return f'<font color="#ef4444"><b>{label}</b></font>'

        for idx, v in enumerate(qs, 1):
            review_display = v.get_review_status_display()
            data.append([
                Paragraph(str(idx), s_cell_center),
                Paragraph(str(v.id), s_cell_center),
                Paragraph(v.detected_at.strftime('%Y-%m-%d'), s_cell_center),
                Paragraph(v.detected_at.strftime('%H:%M:%S'), s_cell_center),
                Paragraph(v.camera.name if v.camera else 'Unknown', s_cell_center),
                Paragraph(v.get_classification_display(), s_cell_center),
                Paragraph(self._get_display_plate_number(v), s_cell_center),
                Paragraph(v.get_detection_status_display(), s_cell_center),
                Paragraph(_conf_color(v.confidence_score), s_cell_center),
                Paragraph(_status_color(review_display), s_cell_center),
            ])

        col_widths = [28, 36, 68, 56, 72, 90, 72, 64, 60, 60]
        table = Table(data, colWidths=col_widths, repeatRows=1)

        style_cmds = [
            # header
            ('BACKGROUND', (0, 0), (-1, 0), BRAND),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('TOPPADDING', (0, 0), (-1, 0), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
            # body
            ('TOPPADDING', (0, 1), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            # grid
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, ACCENT),
            ('LINEBELOW', (0, 1), (-1, -2), 0.5, BORDER),
            ('LINEBELOW', (0, -1), (-1, -1), 1, BRAND),
            ('LINEBEFORE', (0, 0), (0, -1), 0.5, BORDER),
            ('LINEAFTER', (-1, 0), (-1, -1), 0.5, BORDER),
        ]

        # alternating row colours
        for i in range(1, len(data)):
            if i % 2 == 0:
                style_cmds.append(('BACKGROUND', (0, i), (-1, i), LIGHT_BG))

        table.setStyle(TableStyle(style_cmds))
        elements.append(table)

        # ── build ────────────────────────────────────────────────
        doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="saferide_violations_'
            f'{now.strftime("%Y%m%d_%H%M%S")}.pdf"'
        )
        return response
