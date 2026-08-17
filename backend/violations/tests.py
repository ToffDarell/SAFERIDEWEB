import csv
from io import BytesIO
from zipfile import ZipFile
from datetime import timedelta

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase
from rest_framework_api_key.models import APIKey

from cameras.models import Camera
from users.models import AdminNotification, UserNotification, UserProfile, get_default_operator_permissions
from violations.models import Violation


@override_settings(SECURE_SSL_REDIRECT=False)
class ViolationAnalyticsTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="operator",
            email="operator@example.com",
            password="password123",
        )
        UserProfile.objects.update_or_create(
            user=self.user,
            defaults={"role": "tmc_operator", "status": "approved"},
        )
        self.admin_user = User.objects.create_user(
            username="admin",
            email="admin@example.com",
            password="password123",
        )
        self.admin_user.is_staff = True
        self.admin_user.save(update_fields=["is_staff"])
        UserProfile.objects.update_or_create(
            user=self.admin_user,
            defaults={"role": "admin", "status": "approved"},
        )

        self.camera_a = Camera.objects.create(
            name="Camera A",
            location="North Road",
            stream_url="rtsp://example.com/a",
        )
        self.camera_b = Camera.objects.create(
            name="Camera B",
            location="South Road",
            stream_url="rtsp://example.com/b",
        )

        now = timezone.now()
        today = now
        two_days_ago = now - timedelta(days=2)
        six_days_ago = now - timedelta(days=6)
        eight_days_ago = now - timedelta(days=8)

        self.violation_with_evidence = Violation.objects.create(
            camera=self.camera_a,
            detected_at=today,
            detection_status="violation",
            confidence_score=0.91,
            classification="no_helmet",
            plate_number="ABC1234",
            review_status="pending",
            evidence_image=SimpleUploadedFile(
                "violation.gif",
                (
                    b"GIF89a\x01\x00\x01\x00\x80\x00\x00\x00\x00\x00\xff\xff\xff!"
                    b"\xf9\x04\x01\n\x00\x01\x00,\x00\x00\x00\x00\x01\x00\x01\x00"
                    b"\x00\x02\x02L\x01\x00;"
                ),
                content_type="image/gif",
            ),
            plate_crop_image=SimpleUploadedFile(
                "plate.gif",
                (
                    b"GIF89a\x01\x00\x01\x00\x80\x00\x00\x00\x00\x00\xff\xff\xff!"
                    b"\xf9\x04\x01\n\x00\x01\x00,\x00\x00\x00\x00\x01\x00\x01\x00"
                    b"\x00\x02\x02L\x01\x00;"
                ),
                content_type="image/gif",
            ),
        )
        Violation.objects.create(
            camera=self.camera_a,
            detected_at=today,
            detection_status="violation",
            confidence_score=0.83,
            classification="nutshell",
            plate_number="XYZ9876",
            review_status="reviewed",
        )
        Violation.objects.create(
            camera=self.camera_b,
            detected_at=two_days_ago,
            detection_status="violation",
            confidence_score=0.78,
            classification="helmet",
            review_status="pending",
        )
        Violation.objects.create(
            camera=self.camera_b,
            detected_at=six_days_ago,
            detection_status="violation",
            confidence_score=0.88,
            classification="license_plate",
            review_status="resolved",
        )
        Violation.objects.create(
            camera=self.camera_b,
            detected_at=eight_days_ago,
            detection_status="violation",
            confidence_score=0.72,
            classification="no_helmet",
            review_status="pending",
        )

    def _set_plate_correction(self, violation, corrected_plate="845KTH"):
        violation.plate_number_corrected = corrected_plate
        violation.plate_corrected_by = self.admin_user
        violation.plate_corrected_at = timezone.now()
        violation.save(update_fields=[
            "plate_number_corrected",
            "plate_corrected_by",
            "plate_corrected_at",
        ])

    def test_summary_endpoint_returns_full_dataset_counts(self):
        self.client.force_authenticate(user=self.user)

        response = self.client.get("/api/violations/summary/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["total_violations"], 5)
        self.assertEqual(response.data["pending_violations"], 3)
        self.assertEqual(response.data["today_violations"], 2)
        self.assertEqual(response.data["this_week_violations"], 4)

        by_class = {item["classification"]: item["count"] for item in response.data["by_class"]}
        self.assertEqual(by_class["no_helmet"], 2)
        self.assertEqual(by_class["helmet"], 1)
        self.assertEqual(by_class["nutshell"], 1)
        self.assertEqual(by_class["license_plate"], 1)

        by_camera = {item["camera_name"]: item["count"] for item in response.data["by_camera"]}
        self.assertEqual(by_camera["Camera A"], 2)
        self.assertEqual(by_camera["Camera B"], 3)

    def test_weekly_chart_endpoint_returns_last_seven_days(self):
        self.client.force_authenticate(user=self.user)

        response = self.client.get("/api/violations/weekly-chart/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 7)

        counts_by_date = {item["date"]: item["count"] for item in response.data}
        today = timezone.localdate().isoformat()
        two_days_ago = (timezone.localdate() - timedelta(days=2)).isoformat()
        six_days_ago = (timezone.localdate() - timedelta(days=6)).isoformat()
        eight_days_ago = (timezone.localdate() - timedelta(days=8)).isoformat()

        self.assertEqual(counts_by_date[today], 2)
        self.assertEqual(counts_by_date[two_days_ago], 1)
        self.assertEqual(counts_by_date[six_days_ago], 1)
        self.assertNotIn(eight_days_ago, counts_by_date)

    def test_violation_list_supports_specific_date_filter(self):
        self.client.force_authenticate(user=self.user)

        response = self.client.get(
            "/api/violations/",
            {"specific_date": timezone.localdate().isoformat()},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 2)

    def test_summary_supports_specific_month_filter(self):
        Violation.objects.create(
            camera=self.camera_a,
            detected_at=timezone.now() - timedelta(days=40),
            detection_status="violation",
            confidence_score=0.74,
            classification="helmet",
            review_status="pending",
        )
        expected_total = Violation.objects.filter(
            detected_at__year=timezone.localdate().year,
            detected_at__month=timezone.localdate().month,
        ).count()
        self.client.force_authenticate(user=self.user)

        response = self.client.get(
            "/api/violations/summary/",
            {"specific_month": timezone.localdate().strftime("%Y-%m")},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["total_violations"], expected_total)

    def test_weekly_chart_supports_custom_date_range(self):
        self.client.force_authenticate(user=self.user)

        response = self.client.get(
            "/api/violations/weekly-chart/",
            {
                "date_from": (timezone.localdate() - timedelta(days=2)).isoformat(),
                "date_to": timezone.localdate().isoformat(),
            },
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 3)

        counts_by_date = {item["date"]: item["count"] for item in response.data}
        self.assertEqual(counts_by_date[timezone.localdate().isoformat()], 2)
        self.assertEqual(counts_by_date[(timezone.localdate() - timedelta(days=2)).isoformat()], 1)

    def test_operator_violation_detail_returns_full_plate_and_protected_evidence_url(self):
        self.client.force_authenticate(user=self.user)

        response = self.client.get(f"/api/violations/{self.violation_with_evidence.id}/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["plate_number"], "ABC1234")
        self.assertTrue(response.data["has_evidence_image"])
        self.assertTrue(response.data["has_plate_crop_image"])
        self.assertTrue(
            response.data["evidence_image"].endswith(
                f"/api/violations/{self.violation_with_evidence.id}/evidence/"
            )
        )
        self.assertTrue(
            response.data["evidence_download_url"].endswith(
                f"/api/violations/{self.violation_with_evidence.id}/evidence/?download=1"
            )
        )
        self.assertTrue(
            response.data["plate_crop_image"].endswith(
                f"/api/violations/{self.violation_with_evidence.id}/evidence/?variant=plate"
            )
        )
        self.assertTrue(
            response.data["plate_crop_download_url"].endswith(
                f"/api/violations/{self.violation_with_evidence.id}/evidence/?variant=plate&download=1"
            )
        )

    def test_admin_violation_detail_returns_full_plate_number(self):
        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get(f"/api/violations/{self.violation_with_evidence.id}/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["plate_number"], "ABC1234")

    def test_admin_can_correct_plate_number_without_overwriting_original_ocr_value(self):
        self.client.force_authenticate(user=self.admin_user)

        response = self.client.patch(
            f"/api/violations/{self.violation_with_evidence.id}/correct-plate/",
            {"plate_number_corrected": "845KTH"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["plate_number"], "ABC1234")
        self.assertEqual(response.data["plate_number_corrected"], "845KTH")
        self.assertEqual(response.data["plate_corrected_by"], "admin")
        self.assertIsNotNone(response.data["plate_corrected_at"])

        violation = Violation.objects.get(pk=self.violation_with_evidence.pk)
        self.assertEqual(violation.plate_number, "ABC1234")
        self.assertEqual(violation.plate_number_corrected, "845KTH")
        self.assertEqual(violation.plate_corrected_by, self.admin_user)
        self.assertIsNotNone(violation.plate_corrected_at)

    def test_non_admin_cannot_correct_plate_number(self):
        self.client.force_authenticate(user=self.user)

        response = self.client.patch(
            f"/api/violations/{self.violation_with_evidence.id}/correct-plate/",
            {"plate_number_corrected": "845KTH"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_reports_permission_allows_violation_list_but_not_detail_or_evidence(self):
        self.user.profile.permissions = {
            **get_default_operator_permissions(),
            "can_view_violations": False,
            "can_view_reports": True,
        }
        self.user.profile.save(update_fields=["permissions"])
        self.client.force_authenticate(user=self.user)

        list_response = self.client.get("/api/violations/")
        self.assertEqual(list_response.status_code, status.HTTP_200_OK)

        detail_response = self.client.get(f"/api/violations/{self.violation_with_evidence.id}/")
        self.assertEqual(detail_response.status_code, status.HTTP_403_FORBIDDEN)

        evidence_response = self.client.get(f"/api/violations/{self.violation_with_evidence.id}/evidence/")
        self.assertEqual(evidence_response.status_code, status.HTTP_403_FORBIDDEN)

    def test_operator_plate_search_supports_partial_match(self):
        self.client.force_authenticate(user=self.user)

        response = self.client.get("/api/violations/", {"search": "ABC"})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["plate_number"], "ABC1234")
        self.assertTrue(
            AdminNotification.objects.filter(
                notification_type="plate_search",
                actor=self.user,
            ).exists()
        )

    def test_plate_search_requires_at_least_three_characters(self):
        self.client.force_authenticate(user=self.user)

        response = self.client.get("/api/violations/", {"search": "A"})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 0)

    def test_operator_plate_search_rejects_queries_longer_than_three_characters(self):
        self.client.force_authenticate(user=self.user)

        response = self.client.get("/api/violations/", {"search": "ABCD"})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 0)

    def test_admin_plate_search_allows_queries_shorter_than_three_characters(self):
        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/violations/", {"search": "A"})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["plate_number"], "ABC1234")

    def test_operator_without_violation_and_reports_permissions_cannot_list_or_view_analytics(self):
        self.user.profile.permissions = {
            **get_default_operator_permissions(),
            "can_view_violations": False,
            "can_view_reports": False,
        }
        self.user.profile.save(update_fields=["permissions"])
        self.client.force_authenticate(user=self.user)

        list_response = self.client.get("/api/violations/")
        self.assertEqual(list_response.status_code, status.HTTP_403_FORBIDDEN)

        summary_response = self.client.get("/api/violations/summary/")
        self.assertEqual(summary_response.status_code, status.HTTP_403_FORBIDDEN)

        weekly_response = self.client.get("/api/violations/weekly-chart/")
        self.assertEqual(weekly_response.status_code, status.HTTP_403_FORBIDDEN)

    def test_admin_plate_search_filters_by_plate_number_only(self):
        self.client.force_authenticate(user=self.admin_user)

        plate_response = self.client.get("/api/violations/", {"search": "ABC1234"})
        self.assertEqual(plate_response.status_code, status.HTTP_200_OK)
        self.assertEqual(plate_response.data["count"], 1)
        self.assertEqual(plate_response.data["results"][0]["plate_number"], "ABC1234")

        camera_response = self.client.get("/api/violations/", {"search": "Camera A"})
        self.assertEqual(camera_response.status_code, status.HTTP_200_OK)
        self.assertEqual(camera_response.data["count"], 0)

    def test_authenticated_user_can_view_protected_evidence_inline(self):
        self.client.force_authenticate(user=self.user)

        response = self.client.get(f"/api/violations/{self.violation_with_evidence.id}/evidence/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("image/gif", response["Content-Type"])
        self.assertEqual(response["Cache-Control"], "private, no-store")
        self.assertIn("inline", response["Content-Disposition"])
        self.assertTrue(
            AdminNotification.objects.filter(
                notification_type="evidence_view",
                actor=self.user,
                violation=self.violation_with_evidence,
            ).exists()
        )

    def test_authenticated_user_can_download_protected_evidence_as_attachment(self):
        self.client.force_authenticate(user=self.user)

        response = self.client.get(
            f"/api/violations/{self.violation_with_evidence.id}/evidence/",
            {"download": "1"},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("image/gif", response["Content-Type"])
        self.assertEqual(response["Cache-Control"], "private, no-store")
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(".gif", response["Content-Disposition"])

    def test_authenticated_user_can_view_plate_crop_inline(self):
        self.client.force_authenticate(user=self.user)

        response = self.client.get(
            f"/api/violations/{self.violation_with_evidence.id}/evidence/",
            {"variant": "plate"},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("image/gif", response["Content-Type"])
        self.assertEqual(response["Cache-Control"], "private, no-store")
        self.assertIn("inline", response["Content-Disposition"])

    def test_protected_evidence_endpoint_requires_authentication(self):
        response = self.client.get(f"/api/violations/{self.violation_with_evidence.id}/evidence/")

        self.assertIn(response.status_code, {status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN})

    def test_export_endpoint_returns_xlsx_with_tmc_branding_for_operator_with_permission(self):
        self.user.profile.permissions = {
            **get_default_operator_permissions(),
            "can_view_reports": True,
            "can_export_reports": True,
        }
        self.user.profile.save(update_fields=["permissions"])
        self.client.force_authenticate(user=self.user)
        response = self.client.get("/api/violations/export/", {"export_format": "xlsx"})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response['Content-Type'],
        )

        workbook = load_workbook(filename=BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet['A5'].value, 'Traffic Management Center (TMC)')
        self.assertEqual(worksheet['A6'].value, 'SafeRide Violation Report')
        self.assertEqual(worksheet['A14'].value, '#')

        archive = ZipFile(BytesIO(response.content))
        self.assertTrue(any(name.startswith('xl/media/') for name in archive.namelist()))
        self.assertTrue(
            AdminNotification.objects.filter(
                notification_type="report_export",
                actor=self.user,
            ).exists()
        )

    def test_export_endpoint_respects_plate_search_filter(self):
        self.user.profile.permissions = {
            **get_default_operator_permissions(),
            "can_view_reports": True,
            "can_export_reports": True,
        }
        self.user.profile.save(update_fields=["permissions"])
        self._set_plate_correction(self.violation_with_evidence)
        self.client.force_authenticate(user=self.user)

        response = self.client.get(
            "/api/violations/export/",
            {"export_format": "xlsx", "search": "ABC"},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        workbook = load_workbook(filename=BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet["A8"].value, "Total Records: 1")

        exported_plate_numbers = [
            row[6]
            for row in worksheet.iter_rows(min_row=15, values_only=True)
            if row and any(value is not None for value in row)
        ]
        self.assertEqual(exported_plate_numbers, ["845KTH"])

    def test_csv_export_includes_plate_correction_column_and_uses_corrected_value(self):
        self.user.profile.permissions = {
            **get_default_operator_permissions(),
            "can_view_reports": True,
            "can_export_reports": True,
        }
        self.user.profile.save(update_fields=["permissions"])
        self._set_plate_correction(self.violation_with_evidence)
        self.client.force_authenticate(user=self.user)

        response = self.client.get(
            "/api/violations/export/",
            {"export_format": "csv", "search": "ABC"},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        content = response.content.decode("utf-8-sig")
        rows = list(csv.reader(content.splitlines()))

        header_row = next(row for row in rows if row and row[0] == "#")
        self.assertEqual(header_row[6], "Plate Number")
        self.assertEqual(header_row[7], "Plate Corrected")

        data_row = next(row for row in rows if row and len(row) >= 8 and row[0] == "1")
        self.assertEqual(data_row[6], "845KTH")
        self.assertEqual(data_row[7], "Yes")

    def test_export_endpoint_requires_report_view_permission(self):
        self.user.profile.permissions = {
            **get_default_operator_permissions(),
            "can_view_reports": False,
            "can_export_reports": True,
        }
        self.user.profile.save(update_fields=["permissions"])
        self.client.force_authenticate(user=self.user)

        response = self.client.get("/api/violations/export/", {"export_format": "xlsx"})

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_status_update_is_blocked_without_permission(self):
        self.user.profile.permissions = {
            **get_default_operator_permissions(),
            "can_update_violation_status": False,
        }
        self.user.profile.save(update_fields=["permissions"])
        self.client.force_authenticate(user=self.user)

        response = self.client.patch(
            f"/api/violations/{self.violation_with_evidence.id}/",
            {"review_status": "reviewed"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


@override_settings(SECURE_SSL_REDIRECT=False)
class ViolationServiceAuthTests(APITestCase):
    def setUp(self):
        self.operator_user = User.objects.create_user(
            username="service_operator",
            email="service-operator@example.com",
            password="password123",
        )
        UserProfile.objects.update_or_create(
            user=self.operator_user,
            defaults={"role": "tmc_operator", "status": "approved"},
        )
        self.camera = Camera.objects.create(
            name="Service Camera",
            location="South Road",
            stream_url="rtsp://example.com/service",
        )
        _, self.api_key = APIKey.objects.create_key(name="YOLO Service")

    def test_api_key_can_create_violation(self):
        payload = {
            "camera": self.camera.id,
            "detected_at": timezone.now().isoformat(),
            "detection_status": "violation",
            "confidence_score": 0.94,
            "classification": "no_helmet",
            "plate_number": "ABC123",
            "bounding_box": {"x1": 1, "y1": 2, "x2": 3, "y2": 4},
        }

        response = self.client.post(
            "/api/violations/",
            payload,
            format="json",
            HTTP_AUTHORIZATION=f"Api-Key {self.api_key}",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Violation.objects.count(), 1)
        self.assertEqual(Violation.objects.get().classification, "no_helmet")
        created_violation = Violation.objects.get()
        self.assertTrue(
            AdminNotification.objects.filter(
                violation=created_violation,
                notification_type="new_detection",
            ).exists()
        )
        self.assertTrue(
            UserNotification.objects.filter(
                recipient=self.operator_user,
                notification_type="new_detection",
            ).exists()
        )

    def test_api_key_rejects_no_helmet_when_detected_objects_only_show_pedestrian(self):
        payload = {
            "camera": self.camera.id,
            "detected_at": timezone.now().isoformat(),
            "detection_status": "violation",
            "confidence_score": 0.94,
            "classification": "no_helmet",
            "plate_number": "",
            "bounding_box": {"x1": 1, "y1": 2, "x2": 3, "y2": 4},
            "detected_objects": {
                "objects": [
                    {"class": "person"},
                    {"label": "walking"},
                ]
            },
        }

        response = self.client.post(
            "/api/violations/",
            payload,
            format="json",
            HTTP_AUTHORIZATION=f"Api-Key {self.api_key}",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("detected_objects", response.data)
        self.assertEqual(Violation.objects.count(), 0)

    def test_api_key_allows_no_helmet_when_detected_objects_include_vehicle_context(self):
        payload = {
            "camera": self.camera.id,
            "detected_at": timezone.now().isoformat(),
            "detection_status": "violation",
            "confidence_score": 0.94,
            "classification": "no_helmet",
            "plate_number": "",
            "bounding_box": {"x1": 1, "y1": 2, "x2": 3, "y2": 4},
            "detected_objects": {
                "objects": [
                    {"class": "person"},
                    {"class": "motorcycle"},
                    {"class": "no_helmet"},
                ]
            },
        }

        response = self.client.post(
            "/api/violations/",
            payload,
            format="json",
            HTTP_AUTHORIZATION=f"Api-Key {self.api_key}",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Violation.objects.count(), 1)
        self.assertEqual(Violation.objects.get().classification, "no_helmet")

    def test_api_key_can_create_violation_with_multipart_json_fields(self):
        evidence_image = SimpleUploadedFile(
            "violation.gif",
            (
                b"GIF89a\x01\x00\x01\x00\x80\x00\x00\x00\x00\x00\xff\xff\xff!"
                b"\xf9\x04\x01\n\x00\x01\x00,\x00\x00\x00\x00\x01\x00\x01\x00"
                b"\x00\x02\x02L\x01\x00;"
            ),
            content_type="image/gif",
        )
        plate_crop_image = SimpleUploadedFile(
            "plate.gif",
            (
                b"GIF89a\x01\x00\x01\x00\x80\x00\x00\x00\x00\x00\xff\xff\xff!"
                b"\xf9\x04\x01\n\x00\x01\x00,\x00\x00\x00\x00\x01\x00\x01\x00"
                b"\x00\x02\x02L\x01\x00;"
            ),
            content_type="image/gif",
        )
        payload = {
            "camera": str(self.camera.id),
            "detected_at": timezone.now().isoformat(),
            "detection_status": "violation",
            "confidence_score": "0.77",
            "classification": "nutshell",
            "plate_number": "",
            "bounding_box": "{'x1': 11, 'y1': 22, 'x2': 33, 'y2': 44}",
            "evidence_image": evidence_image,
            "plate_crop_image": plate_crop_image,
        }

        response = self.client.post(
            "/api/violations/",
            payload,
            format="multipart",
            HTTP_AUTHORIZATION=f"Api-Key {self.api_key}",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Violation.objects.count(), 1)
        violation = Violation.objects.get()
        self.assertEqual(violation.classification, "nutshell")
        self.assertEqual(violation.bounding_box, {"x1": 11, "y1": 22, "x2": 33, "y2": 44})
        self.assertTrue(bool(violation.plate_crop_image))
